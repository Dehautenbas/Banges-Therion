# =========================
# installer cette librairie

# pip install pandas openpyxl
# =========================


import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# PARAMÈTRES
# =========================
# SQL_DUMP_PATH = "E9_glaciere-du-furggeli/Exports/sardines_database24.sql"
# OUTPUT_XLSX = "E9_glaciere-du-furggeli/Exports/analyse-database_glacière.xlsx"

SQL_DUMP_PATH = "/Users/martinkern/Documents/GitHub/Windgallen-Therion/datas/F1_sardines/Exports/sardines_database.sql"
OUTPUT_XLSX = "/Users/martinkern/Documents/GitHub/Windgallen-Therion/datas/F1_sardines/Exports/analyse-database_sardines.xlsx"

# =========================
# 1. Charger la base SQL
# =========================
conn = sqlite3.connect(":memory:")
cur = conn.cursor()

with open(SQL_DUMP_PATH, "r", encoding="utf-8") as f:
    cur.executescript(f.read())

# =========================
# 2. Agrégats par survey
# =========================
agg_query = """
SELECT
    s.ID AS survey_id,
    COUNT(DISTINCT c.ID) AS nb_centrelines,
    COALESCE(SUM(c.LENGTH), 0) AS total_length
FROM SURVEY s
LEFT JOIN CENTRELINE c ON c.SURVEY_ID = s.ID
GROUP BY s.ID
"""

# =========================
# 3. Détail + synthèse
# =========================
query = """
WITH persons_link AS (
    SELECT CENTRELINE_ID, PERSON_ID FROM EXPLO
    UNION
    SELECT CENTRELINE_ID, PERSON_ID FROM TOPO
),
survey_length AS (
    SELECT
        s.ID AS survey_id,
        ROUND(COALESCE(SUM(c.LENGTH),0),2) AS longueur_survey
    FROM SURVEY s
    LEFT JOIN CENTRELINE c ON c.SURVEY_ID = s.ID
    GROUP BY s.ID
)

SELECT DISTINCT
    s.NAME AS survey_name,
    c.EXPLO_DATE,
    p.NAME AS person_name,
    p.SURNAME AS person_surname,
    sl.longueur_survey
FROM SURVEY s
LEFT JOIN survey_length sl ON sl.survey_id = s.ID
LEFT JOIN CENTRELINE c ON c.SURVEY_ID = s.ID
LEFT JOIN persons_link pl ON pl.CENTRELINE_ID = c.ID
LEFT JOIN PERSON p ON p.ID = pl.PERSON_ID
WHERE s.NAME NOT LIKE '%_tot'
AND (c.EXPLO_DATE IS NOT NULL OR p.ID IS NOT NULL)
ORDER BY
    c.EXPLO_DATE,
    s.NAME,
    p.SURNAME

"""
df = pd.read_sql_query(query, conn)

# =========================
# 4. Ligne vide entre surveys
# =========================
rows = []
last_survey = None

for _, row in df.iterrows():
    current_survey = row["survey_name"]

    if last_survey is not None and current_survey != last_survey:
        rows.append([None] * len(df.columns))  # ligne vide

    rows.append(row.tolist())
    last_survey = current_survey

df = pd.DataFrame(rows, columns=df.columns)

# =========================
# 5. Export Excel brut
# =========================
df.to_excel(OUTPUT_XLSX, index=False)

# =========================
# 6. Calculs globaux
# =========================
total_length = conn.execute("""
SELECT ROUND(COALESCE(SUM(LENGTH),0),2)
FROM CENTRELINE
""").fetchone()[0]

total_surveys = conn.execute("""
SELECT COUNT(DISTINCT ID)
FROM SURVEY
WHERE NAME NOT LIKE '%_tot'
""").fetchone()[0]

total_centrelines = conn.execute("""
SELECT COUNT(DISTINCT ID)
FROM CENTRELINE
""").fetchone()[0]

year_totals = conn.execute("""
SELECT
    SUBSTR(EXPLO_DATE,1,4) AS year,
    ROUND(SUM(LENGTH),2)
FROM CENTRELINE
WHERE EXPLO_DATE IS NOT NULL
GROUP BY year
ORDER BY year
""").fetchall()

# =========================
# 7. Mise en forme Excel
# =========================
wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

# Styles
header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
survey_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
section_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
bold = Font(bold=True)
center = Alignment(vertical="center")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# En-têtes
for col in range(1, ws.max_column + 1):
    c = ws.cell(row=1, column=col)
    c.font = bold
    c.fill = header_fill
    c.border = border
    ws.column_dimensions[get_column_letter(col)].width = 22

# Bordures & alignement
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = center

# =========================
# 8. Regroupement visuel par survey
# =========================
survey_col = 1
last_survey = None
row = 2

while row <= ws.max_row:
    val = ws.cell(row=row, column=survey_col).value
    if val and val != last_survey:
        ws.insert_rows(row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
        cell = ws.cell(row=row, column=1, value=val)
        cell.font = bold
        cell.fill = survey_fill
        last_survey = val
        row += 1
    row += 1

# =========================
# 9. Bloc TOTALS
# =========================
start_row = ws.max_row + 2

ws.cell(row=start_row, column=1, value="TOTAL GLOBAL").font = bold
ws.cell(row=start_row, column=1).fill = section_fill
ws.cell(row=start_row, column=2, value=f"{total_length:.2f} m").font = bold
ws.cell(row=start_row, column=2).fill = section_fill

ws.cell(row=start_row + 1, column=1, value="Nombre total de surveys").font = bold
ws.cell(row=start_row + 1, column=2, value=total_surveys).font = bold

ws.cell(row=start_row + 2, column=1, value="Nombre total de centrelines").font = bold
ws.cell(row=start_row + 2, column=2, value=total_centrelines).font = bold

# Bloc année
row_year = start_row + 4
ws.cell(row=row_year, column=1, value="TOTAL PAR ANNÉE").font = bold
ws.cell(row=row_year, column=1).fill = section_fill

row_year += 1

for year, total in year_totals:
    ws.cell(row=row_year, column=1, value=year).font = bold
    ws.cell(row=row_year, column=2, value=round(total, 2)).font = bold
    row_year += 1

wb.save(OUTPUT_XLSX)

print(f"✅ Fichier généré : {OUTPUT_XLSX}")
